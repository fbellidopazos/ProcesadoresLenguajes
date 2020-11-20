from openpyxl import load_workbook,workbook,cell


workbook = load_workbook(filename="Sint_Ascendente.xlsx")
#print(workbook.sheetnames)

"""
Tabla Accion
"""

sheet = workbook["Accion"]

file = open("toJavaAccion.txt","w")


file.write("Pair<String, Integer>[][] tablaAccion=new Pair<>[105][29];\n")
for i in range(2,107):
    for j in range(2,32):

        content = sheet.cell(row=i, column=j).value
        if(content != None):
            if(content[0]=="s"):
                file.write(f"tablaAccion[{i-2}][{j-2}] = new Pair<>(\"S\", {(int)(content[1::])});\n")
            elif(content[0]=="r"):
                file.write(f"tablaAccion[{i-2}][{j-2}] = new Pair<>(\"R\", {(int)(content[1::])});\n")
            else:
                file.write(f"tablaAccion[{i-2}][{j-2}] = new Pair<>(\"ACC\", null);\n")
file.close()       
        

file = open("toJavaAccionHashMap.txt","w")
file.write("HashMap<String, Integer> aplicacionTerminal = new HashMap<>();\n")
for j in range(1,32):
    content = sheet.cell(row=1, column=j).value
    if(content != None):
            file.write(f"aplicacionTerminal.put(\"{content}\", {j-2});\n")



file.close()
"""
Tabla GoTo
"""

sheet = workbook["GoTo"]

file = open("toJavaGoTo.txt","w")


file.write("int[][] tablaGoTo = new int[105][21];\n")
for i in range(2,107):
    for j in range(1,22):

        content = sheet.cell(row=i, column=j).value
        if(content != None):
            file.write(f'tablaGoTo[{i-2}][{j-1}] = {(int)(content)};\n')
file.close()


file = open("toJavaGoToHashMap.txt","w")
file.write("HashMap<String, Integer> aplicacionNoTerminal = new HashMap<>();\n")
for j in range(1,32):
    content = sheet.cell(row=1, column=j).value
    if(content != None):
            file.write(f"aplicacionNoTerminal.put(\"{content}\", {j-1});\n")

file.close()



'''
https://realpython.com/openpyxl-excel-spreadsheets-python/
'''