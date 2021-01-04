from openpyxl import load_workbook,workbook,cell


workbook = load_workbook(filename="Sint_Ascendente.xlsx")
#print(workbook.sheetnames)

"""
Tabla Accion
"""

sheet = workbook["Accion"]

file = open("toJavaAccion.txt","w")
file2 = open("Errores.txt","w")

file.write("Pair<String, Integer>[][] tablaAccion=new Pair[105][29];\n")
for i in range(2,107):
    for j in range(2,31):

        content = sheet.cell(row=i, column=j).value
        if(content != None):
            if(content[0]=="s"):
                file.write(f"tablaAccion[{i-2}][{j-2}] = new Pair<>(\"S\", {(int)(content[1::])});\n")
            elif(content[0]=="r"):
                file.write(f"tablaAccion[{i-2}][{j-2}] = new Pair<>(\"R\", {(int)(content[1::])});\n")
            else:
                file.write(f"tablaAccion[{i-2}][{j-2}] = new Pair<>(\"ACC\", null);\n")
        else:
            k=i-2
            firsts={
                "Z":"if,let,identificador,alert,input,return,do,function o nada",
                "P":"if,let,identificador,alert,input,return,do,function o nada",
                "F":"function",
                "I":"function",
                "J":"(",
                "G":"{",
                "H":"number,boolean,string o nada",
                "A":"number,boolean,string o nada",
                "K":", o nada",
                "C":"if,let,identificador,alert,input,return,do o nada",
                "S":"identificador,alert,input,return",
                "B":"if,let,identificador,alert,input,return,do",
                "T":"number,boolean,string",
                "E":"identificador,( ,cteEntera,cadena",
                "R":"identificador,( ,cteEntera,cadena",
                "U":"identificador,( ,cteEntera,cadena",
                "V":"identificador,( ,cteEntera,cadena",
                "W":"identificador,( ,cteEntera,cadena",
                "L":"identificador,( ,cteEntera,cadena o nada",
                "Q":", o nada",
                "X":"identificador,( ,cteEntera,cadena o nada"
            }

            switch = {
                0: "Se esperaba:"+firsts.get("P"),
                1: "Se esperaba: Fin de Programa",
                2: "Se esperaba: "+firsts.get("P"),
                3: "Se esperaba: "+firsts.get("P"),
                4: "Se esperaba: (",
                5: "Se esperaba: "+firsts.get("T"),
                6: "No se esperaba nada",
                7: "Se esperaba: {",
                8: "Se esperaba: "+firsts.get("J"),
                9: "Se esperaba: =",
                10: "Se esperaba: (",
                11: "Se esperaba: (",
                12: "Se esperaba: "+firsts.get("X"),
                13: "Se esperaba: "+firsts.get("H"),
                14: "No se esperaba nada",
                15: "No se esperaba nada",
                16: "Se esperaba: "+firsts.get("E"),
                17: "Se esperaba: identificador",
                18: "No se esperaba nada",
                19: "No se esperaba nada",
                20: "No se esperaba nada",
                21: "Se esperaba: "+firsts.get("C"),
                22: "Se esperaba: "+firsts.get("G"),
                23: "Se esperaba: "+firsts.get("A"),
                24: "Se esperaba: "+firsts.get("E"),
                25: "Se esperaba: "+firsts.get("L"),
                26: "Se esperaba: "+firsts.get("E"),
                27: "Se esperaba: "+firsts.get("E"),
                28: "Se esperaba: identificador",
                29: "Se esperaba: ;",
                30: "Se esperaba: ||",
                31: "Se esperaba: &&",
                32: "Se esperaba: == , !=",
                33: "Se esperaba: +, -",
                34: "No se esperaba nada",
                35: "Se esperaba: (,||,&&,==,!=,+,- o nada",
                36: "Se esperaba: "+firsts.get("E"),
                37: "No se esperaba nada",
                38: "No se esperaba nada",
                39: "Se esperaba: identificador",
                40: "No se esperaba nada",
                41: "Se esperaba: ), ||",
                42: "Se esperaba: ;",
                43: "Se esperaba: }",
                44: "Se esperaba: "+firsts.get("C"),
                45: "No se esperaba nada",
                46: "Se esperaba: "+firsts.get("C"),
                47: "Se esperaba: )",
                48: "Se esperaba: identificador",
                49: "Se esperaba: ;",
                50: "Se esperaba: )",
                51: "Se esperaba: ||, "+firsts.get("Q"),
                52: "Se esperaba: ;, ||",
                53: "Se esperaba: ), ||",
                54: "Se esperaba: )",
                55: "No se esperaba nada",
                56: "Se esperaba: "+firsts.get("R"),
                57: "Se esperaba: "+firsts.get("U"),
                58: "Se esperaba: "+firsts.get("V"),
                59: "Se esperaba: "+firsts.get("V"),
                60: "Se esperaba: "+firsts.get("W"),
                61: "Se esperaba: "+firsts.get("W"),
                62: "Se esperaba: "+firsts.get("L"),
                63: "Se esperaba: ), ||",
                64: "No se esperaba nada",
                65: "Se esperaba: "+firsts.get("S"),
                66: "No se esperaba nada",
                67: "Se esperaba: While",
                68: "No se esperaba nada",
                69: "Se esperaba: }",
                70: "No se esperaba nada",
                71: "Se esperaba: "+firsts.get("K"),
                72: "No se esperaba nada",
                73: "Se esperaba: ;",
                74: "No se esperaba nada",
                75: "Se esperaba: "+firsts.get("E"),
                76: "No se esperaba nada",
                77: "Se esperaba: ;",
                78: "Se esperaba: ;",
                79: "Se esperaba: &&",
                80: "Se esperaba: ==, !=",
                81: "Se esperaba: +, -",
                82: "Se esperaba: +, -",
                83: "No se esperaba nada",
                84: "No se esperaba nada",
                85: "Se esperaba: )",
                86: "No se esperaba nada",
                87: "No se esperaba nada",
                88: "Se esperaba: (",
                89: "No se esperaba nada",
                90: "No se esperaba nada",
                91: "Se esperaba: "+firsts.get("T"),
                92: "No se esperaba nada",
                93: "Se esperaba: ||, "+firsts.get("Q"),
                94: "No se esperaba nada",
                95: "No se esperaba nada",
                96: "No se esperaba nada",
                97: "Se esperaba: "+firsts.get("E"),
                98: "Se esperaba: identificador",
                99: "No se esperaba nada",
                100: "Se esperaba: ) o ||",
                101: "Se esperaba: "+firsts.get("K"),
                102: "Se esperaba: ;",
                103: "No se esperaba nada",
                104: "No se esperaba nada"
            }




            file2.write(f"tablaAccion[{i-2}][{j-2}] = new Pair<>(\"E\", \"{switch.get(k)}\");\n")
file.close()       
file2.close()
        

file = open("toJavaAccionHashMap.txt","w")
file.write("HashMap<String, Integer> aplicacionTerminal = new HashMap<>();\n")
for j in range(1,32):
    content = sheet.cell(row=1, column=j).value
    if(content != None):
            file.write(f"aplicacionTerminal.put(\"{content}\", {j-2});\n")



file.close()
"""
Tabla GoTo
"""

sheet = workbook["GoTo"]

file = open("toJavaGoTo.txt","w")


file.write("int[][] tablaGoTo = new int[105][21];\n")
for i in range(2,150):
    for j in range(1,40):

        content = sheet.cell(row=i, column=j).value
        if(content != None):
            file.write(f'tablaGoTo[{i-2}][{j-1}] = {(int)(content)};\n')
file.close()


file = open("toJavaGoToHashMap.txt","w")
file.write("HashMap<String, Integer> aplicacionNoTerminal = new HashMap<>();\n")
for j in range(1,32):
    content = sheet.cell(row=1, column=j).value
    if(content != None):
            file.write(f"aplicacionNoTerminal.put(\"{content}\", {j-1});\n")

file.close()



'''
https://realpython.com/openpyxl-excel-spreadsheets-python/
'''